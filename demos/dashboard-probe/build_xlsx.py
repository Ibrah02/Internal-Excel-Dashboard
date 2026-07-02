"""Option A probe: a real .xlsx dashboard built the spreadsheet way.
KPIs as live formulas, native charts, a summary table. Opens in LibreOffice/OnlyOffice/Excel.
"""
import csv
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- Shared sample dataset: monthly 2025 sales -----------------------------
regions = ["North", "South", "East", "West"]
products = ["Widgets", "Gadgets", "Gizmos"]
months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# deterministic, realistic-looking numbers (no randomness so both probes match)
base = {"Widgets": 1200, "Gadgets": 800, "Gizmos": 450}
region_mult = {"North": 1.30, "South": 0.85, "East": 1.05, "West": 0.95}
price = {"Widgets": 25, "Gadgets": 60, "Gizmos": 120}

rows = []  # date, region, product, units, revenue
for mi, m in enumerate(months):
    season = 1.0 + 0.28 * ((mi % 6) / 5.0) + (0.15 if m in ("Nov", "Dec") else 0)
    for r in regions:
        for p in products:
            units = round(base[p] * region_mult[r] * season / 10)
            revenue = units * price[p]
            rows.append([f"2025-{mi+1:02d}", r, p, units, revenue])

# write shared CSV (used by Option B too)
with open("sample_sales.csv", "w", newline="") as f:
    w = csv.writer(f)
    w.writerow(["Month", "Region", "Product", "Units", "Revenue"])
    w.writerows(rows)

# --- Build workbook --------------------------------------------------------
wb = Workbook()

# 1) Raw data sheet
ws = wb.active
ws.title = "Data"
hdr = ["Month", "Region", "Product", "Units", "Revenue"]
ws.append(hdr)
for row in rows:
    ws.append(row)
head_fill = PatternFill("solid", fgColor="1D9E75")
for c in range(1, 6):
    cell = ws.cell(1, c)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = head_fill
for r in range(2, len(rows) + 2):
    ws.cell(r, 5).number_format = '#,##0'
ws.freeze_panes = "A2"
for i, wdt in enumerate([10, 10, 12, 10, 14], start=1):
    ws.column_dimensions[get_column_letter(i)].width = wdt
n = len(rows)

# 2) Monthly summary (formulas -> live recalculation like Excel)
mon = wb.create_sheet("Monthly")
mon.append(["Month", "Revenue", "Units"])
for i, m in enumerate(months):
    key = f"2025-{i+1:02d}"
    mon.append([
        m,
        f'=SUMIF(Data!$A$2:$A${n+1},"{key}",Data!$E$2:$E${n+1})',
        f'=SUMIF(Data!$A$2:$A${n+1},"{key}",Data!$D$2:$D${n+1})',
    ])
for c in (1, 2, 3):
    mon.cell(1, c).font = Font(bold=True)
for r in range(2, 14):
    mon.cell(r, 2).number_format = '#,##0'

# 3) By-region and by-product summaries
byreg = wb.create_sheet("ByRegion")
byreg.append(["Region", "Revenue"])
for r in regions:
    byreg.append([r, f'=SUMIF(Data!$B$2:$B${n+1},"{r}",Data!$E$2:$E${n+1})'])
for c in (1, 2):
    byreg.cell(1, c).font = Font(bold=True)

byprod = wb.create_sheet("ByProduct")
byprod.append(["Product", "Revenue"])
for p in products:
    byprod.append([p, f'=SUMIF(Data!$C$2:$C${n+1},"{p}",Data!$E$2:$E${n+1})'])
for c in (1, 2):
    byprod.cell(1, c).font = Font(bold=True)

# 4) Dashboard sheet: KPI tiles + charts
dash = wb.create_sheet("Dashboard", 0)
dash.sheet_view.showGridLines = False
title = dash.cell(1, 1, "2025 Sales Dashboard")
title.font = Font(size=20, bold=True, color="0F6E56")

# KPI tiles as live formulas
kpis = [
    ("Total Revenue", f'=SUM(Data!E2:E{n+1})', '#,##0'),
    ("Total Units", f'=SUM(Data!D2:D{n+1})', '#,##0'),
    ("Avg Order Value", f'=SUM(Data!E2:E{n+1})/SUM(Data!D2:D{n+1})', '#,##0.00'),
    ("Top Region", '=INDEX(ByRegion!A2:A5,MATCH(MAX(ByRegion!B2:B5),ByRegion!B2:B5,0))', '@'),
]
tile_fill = PatternFill("solid", fgColor="E1F5EE")
thin = Side(style="thin", color="9FE1CB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
col = 1
for label, formula, fmt in kpis:
    lab = dash.cell(3, col, label)
    lab.font = Font(size=11, color="0F6E56", bold=True)
    lab.fill = tile_fill
    lab.border = border
    lab.alignment = Alignment(horizontal="left")
    val = dash.cell(4, col, formula)
    val.font = Font(size=18, bold=True)
    val.number_format = fmt
    val.fill = tile_fill
    val.border = border
    dash.column_dimensions[get_column_letter(col)].width = 20
    col += 2

# Chart 1: monthly revenue line
line = LineChart()
line.title = "Revenue by Month"
line.height, line.width = 7.5, 15
data = Reference(mon, min_col=2, min_row=1, max_row=13)
cats = Reference(mon, min_col=1, min_row=2, max_row=13)
line.add_data(data, titles_from_data=True)
line.set_categories(cats)
dash.add_chart(line, "A7")

# Chart 2: revenue by region bar
bar = BarChart()
bar.title = "Revenue by Region"
bar.height, bar.width = 7.5, 8
bdata = Reference(byreg, min_col=2, min_row=1, max_row=5)
bcats = Reference(byreg, min_col=1, min_row=2, max_row=5)
bar.add_data(bdata, titles_from_data=True)
bar.set_categories(bcats)
dash.add_chart(bar, "A24")

# Chart 3: revenue by product pie
pie = PieChart()
pie.title = "Revenue by Product"
pie.height, pie.width = 7.5, 8
pdata = Reference(byprod, min_col=2, min_row=1, max_row=4)
pcats = Reference(byprod, min_col=1, min_row=2, max_row=4)
pie.add_data(pdata, titles_from_data=True)
pie.set_categories(pcats)
dash.add_chart(pie, "F24")

wb.save("SalesDashboard.xlsx")
print("wrote SalesDashboard.xlsx and sample_sales.csv  (", n, "data rows )")
